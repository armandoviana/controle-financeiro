from flask import Flask, render_template, request, jsonify, session, redirect, url_for
import sqlite3
from datetime import datetime, timedelta
from database import init_db
from functools import wraps
import secrets
import hashlib
import os

app = Flask(__name__)
# Gerar chave secreta aleatória e segura
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))
app.config['SESSION_COOKIE_SECURE'] = False  # Mude para True se usar HTTPS
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)

# Inicializar banco de dados
try:
    init_db()
    print("✅ Banco de dados inicializado com sucesso!")
except Exception as e:
    print(f"⚠️ Erro ao inicializar banco: {e}")
    print("Tentando criar tabelas...")

# Inicializar SQLAlchemy (novo)
try:
    from config import Config
    from models import db
    from models.user import User
    from models.transaction import Receita, Gasto
    from models.meta import Meta
    
    app.config.from_object(Config)
    db.init_app(app)
    print("✅ SQLAlchemy inicializado!")
except Exception as e:
    print(f"⚠️ SQLAlchemy não inicializado: {e}")

# Credenciais com hash (MUDE ISSO!)
USUARIO = 'casal'
# Senha: Rebily1234 (hash SHA-256)
SENHA_HASH = 'a47e6dc656c2c1a7adc1197cf93a3401935154611ad1d8f36f8d2158a172e095'

# Limite de tentativas de login
login_attempts = {}
MAX_ATTEMPTS = 5
LOCKOUT_TIME = 300  # 5 minutos

def hash_senha(senha):
    return hashlib.sha256(senha.encode()).hexdigest()

def verificar_bloqueio(ip):
    if ip in login_attempts:
        attempts, last_attempt = login_attempts[ip]
        if attempts >= MAX_ATTEMPTS:
            if datetime.now().timestamp() - last_attempt < LOCKOUT_TIME:
                return True
            else:
                # Reset após tempo de bloqueio
                del login_attempts[ip]
    return False

def registrar_tentativa(ip, sucesso):
    if sucesso:
        if ip in login_attempts:
            del login_attempts[ip]
    else:
        if ip in login_attempts:
            attempts, _ = login_attempts[ip]
            login_attempts[ip] = (attempts + 1, datetime.now().timestamp())
        else:
            login_attempts[ip] = (1, datetime.now().timestamp())

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def api_error_handler(f):
    """Decorator que captura erros e retorna JSON"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        try:
            return f(*args, **kwargs)
        except Exception as e:
            print(f"❌ Erro em {f.__name__}: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({'success': False, 'message': f'Erro: {str(e)}'}), 500
    return decorated_function

def get_db():
    """Retorna conexão com banco de dados (SQLite ou PostgreSQL)"""
    database_url = os.environ.get('DATABASE_URL')
    
    if database_url:
        # PostgreSQL (produção)
        import psycopg2
        import psycopg2.extras
        
        # Render usa postgres://, mas psycopg2 precisa de postgresql://
        if database_url.startswith('postgres://'):
            database_url = database_url.replace('postgres://', 'postgresql://', 1)
        
        conn = psycopg2.connect(database_url, cursor_factory=psycopg2.extras.RealDictCursor)
        return conn
    else:
        # SQLite (local)
        conn = sqlite3.connect('financas.db')
        conn.row_factory = sqlite3.Row
        return conn

def db_execute(conn, query, params=None):
    """Executa query compatível com PostgreSQL e SQLite"""
    is_postgres = hasattr(conn, 'cursor_factory')
    if is_postgres:
        query = query.replace('?', '%s')
    
    cursor = conn.cursor()
    if params:
        cursor.execute(query, params)
    else:
        cursor.execute(query)
    return cursor

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        ip = request.remote_addr
        
        # Verificar se está bloqueado
        if verificar_bloqueio(ip):
            return jsonify({
                'success': False, 
                'message': f'Muitas tentativas. Tente novamente em 5 minutos.'
            }), 429
        
        data = request.json
        usuario = data.get('usuario', '')
        senha = data.get('senha', '')
        
        # Validação básica
        if not usuario or not senha:
            return jsonify({'success': False, 'message': 'Preencha todos os campos'}), 400
        
        senha_hash = hash_senha(senha)
        
        # Buscar usuário no banco
        try:
            conn = get_db()
            cursor = conn.cursor()
            # PostgreSQL usa %s, SQLite usa ?
            query = 'SELECT id, username, password_hash FROM usuarios WHERE username = %s' if hasattr(conn, 'cursor_factory') else 'SELECT id, username, password_hash FROM usuarios WHERE username = ?'
            cursor.execute(query, (usuario,))
            user = cursor.fetchone()
            conn.close()
            
            if user and user['password_hash'] == senha_hash:
                session.permanent = True
                session['logged_in'] = True
                session['user_id'] = user['id']
                session['usuario'] = user['username']
                registrar_tentativa(ip, True)
                return jsonify({'success': True})
            
            registrar_tentativa(ip, False)
            tentativas_restantes = MAX_ATTEMPTS - login_attempts.get(ip, (0, 0))[0]
            return jsonify({
                'success': False, 
                'message': f'Usuário ou senha incorretos. {tentativas_restantes} tentativas restantes.'
            }), 401
        except Exception as e:
            return jsonify({
                'success': False,
                'message': f'Erro no banco de dados. Acesse /api/migrar-db primeiro. Erro: {str(e)}'
            }), 500
    
    return render_template('login.html')

@app.route('/cadastro', methods=['GET', 'POST'])
def cadastro():
    if request.method == 'POST':
        try:
            data = request.json
            username = data.get('username', '').strip()
            email = data.get('email', '').strip()
            senha = data.get('senha', '')
            confirmar_senha = data.get('confirmar_senha', '')
            
            # Validações
            if not username or not senha:
                return jsonify({'success': False, 'message': 'Username e senha são obrigatórios'}), 400
            
            if len(username) < 3:
                return jsonify({'success': False, 'message': 'Username deve ter no mínimo 3 caracteres'}), 400
            
            if len(senha) < 6:
                return jsonify({'success': False, 'message': 'Senha deve ter no mínimo 6 caracteres'}), 400
            
            if senha != confirmar_senha:
                return jsonify({'success': False, 'message': 'As senhas não coincidem'}), 400
            
            # Verificar se username já existe
            conn = get_db()
            cursor = conn.cursor()
            is_postgres = hasattr(conn, 'cursor_factory')
            
            query = 'SELECT id FROM usuarios WHERE username = %s' if is_postgres else 'SELECT id FROM usuarios WHERE username = ?'
            cursor.execute(query, (username,))
            if cursor.fetchone():
                conn.close()
                return jsonify({'success': False, 'message': 'Username já está em uso'}), 400
            
            # Criar usuário
            from datetime import datetime
            senha_hash = hash_senha(senha)
            data_criacao = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            query = 'INSERT INTO usuarios (username, password_hash, email, data_criacao) VALUES (%s, %s, %s, %s)' if is_postgres else 'INSERT INTO usuarios (username, password_hash, email, data_criacao) VALUES (?, ?, ?, ?)'
            cursor.execute(query, (username, senha_hash, email, data_criacao))
            conn.commit()
            conn.close()
            
            return jsonify({'success': True, 'message': 'Conta criada com sucesso!'})
        except Exception as e:
            return jsonify({'success': False, 'message': f'Erro ao criar conta: {str(e)}'}), 500
    
    return render_template('cadastro.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    return render_template('index.html')

@app.route('/api/receitas', methods=['GET', 'POST', 'PUT', 'DELETE'])
@login_required
def receitas():
    conn = get_db()
    user_id = session.get('user_id')
    
    if request.method == 'POST':
        data = request.json
        
        # Validação de dados
        if not data.get('descricao') or not data.get('valor') or not data.get('tipo') or not data.get('data'):
            conn.close()
            return jsonify({'success': False, 'message': 'Dados incompletos'}), 400
        
        try:
            valor = float(data['valor'])
            if valor <= 0:
                raise ValueError('Valor deve ser positivo')
        except (ValueError, TypeError):
            conn.close()
            return jsonify({'success': False, 'message': 'Valor inválido'}), 400
        
        db_execute(conn, 'INSERT INTO receitas (user_id, descricao, valor, tipo, data, notas, tags) VALUES (?, ?, ?, ?, ?, ?, ?)',
                    (user_id, data['descricao'][:200], valor, data['tipo'], data['data'],
                     data.get('notas', ''), data.get('tags', '')))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    
    elif request.method == 'PUT':
        data = request.json
        receita_id = data.get('id')
        
        if not receita_id:
            conn.close()
            return jsonify({'success': False, 'message': 'ID não fornecido'}), 400
        
        try:
            valor = float(data['valor'])
            if valor <= 0:
                raise ValueError('Valor deve ser positivo')
        except (ValueError, TypeError):
            conn.close()
            return jsonify({'success': False, 'message': 'Valor inválido'}), 400
        
        db_execute(conn, 'UPDATE receitas SET descricao=?, valor=?, tipo=?, data=?, notas=?, tags=? WHERE id=? AND user_id=?',
                    (data['descricao'][:200], valor, data['tipo'], data['data'],
                     data.get('notas', ''), data.get('tags', ''), receita_id, user_id))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    
    elif request.method == 'DELETE':
        receita_id = request.args.get('id')
        if not receita_id:
            conn.close()
            return jsonify({'success': False, 'message': 'ID não fornecido'}), 400
        
        db_execute(conn, 'DELETE FROM receitas WHERE id=? AND user_id=?', (receita_id, user_id))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    
    receitas = db_execute(conn, 'SELECT * FROM receitas WHERE user_id=? ORDER BY data DESC', (user_id,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in receitas])

@app.route('/api/gastos', methods=['GET', 'POST', 'PUT', 'DELETE'])
@login_required
def gastos():
    conn = get_db()
    user_id = session.get('user_id')
    
    if request.method == 'POST':
        data = request.json
        
        # Validação de dados
        if not data.get('descricao') or not data.get('valor') or not data.get('categoria') or not data.get('data'):
            conn.close()
            return jsonify({'success': False, 'message': 'Dados incompletos'}), 400
        
        try:
            valor = float(data['valor'])
            if valor <= 0:
                raise ValueError('Valor deve ser positivo')
        except (ValueError, TypeError):
            conn.close()
            return jsonify({'success': False, 'message': 'Valor inválido'}), 400
        
        cursor = db_execute(conn, 'INSERT INTO gastos (user_id, descricao, valor, categoria, data, notas, tags) VALUES (?, ?, ?, ?, ?, ?, ?)',
                    (user_id, data['descricao'][:200], valor, data['categoria'], data['data'], 
                     data.get('notas', ''), data.get('tags', '')))
        gasto_id = cursor.lastrowid
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'id': gasto_id})
    
    elif request.method == 'PUT':
        data = request.json
        gasto_id = data.get('id')
        
        if not gasto_id:
            conn.close()
            return jsonify({'success': False, 'message': 'ID não fornecido'}), 400
        
        try:
            valor = float(data['valor'])
            if valor <= 0:
                raise ValueError('Valor deve ser positivo')
        except (ValueError, TypeError):
            conn.close()
            return jsonify({'success': False, 'message': 'Valor inválido'}), 400
        
        db_execute(conn, 'UPDATE gastos SET descricao=?, valor=?, categoria=?, data=?, notas=?, tags=? WHERE id=? AND user_id=?',
                    (data['descricao'][:200], valor, data['categoria'], data['data'],
                     data.get('notas', ''), data.get('tags', ''), gasto_id, user_id))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    
    elif request.method == 'DELETE':
        gasto_id = request.args.get('id')
        if not gasto_id:
            conn.close()
            return jsonify({'success': False, 'message': 'ID não fornecido'}), 400
        
        db_execute(conn, 'DELETE FROM gastos WHERE id=? AND user_id=?', (gasto_id, user_id))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    
    gastos = db_execute(conn, 'SELECT * FROM gastos WHERE user_id=? ORDER BY data DESC', (user_id,)).fetchall()
    conn.close()
    return jsonify([dict(g) for g in gastos])

@app.route('/api/resumo')
@api_error_handler
@login_required
def resumo():
    conn = get_db()
    user_id = session.get('user_id')
    total_receitas = db_execute(conn, 'SELECT SUM(valor) as total FROM receitas WHERE user_id=?', (user_id,)).fetchone()['total'] or 0
    total_gastos = db_execute(conn, 'SELECT SUM(valor) as total FROM gastos WHERE user_id=?', (user_id,)).fetchone()['total'] or 0
    conn.close()
    return jsonify({
        'receitas': total_receitas,
        'gastos': total_gastos,
        'saldo': total_receitas - total_gastos
    })

@app.route('/api/evolucao')
@api_error_handler
@login_required
def evolucao():
    conn = get_db()
    user_id = session.get('user_id')
    
    # Últimos 6 meses
    try:
        query = '''
            SELECT strftime('%Y-%m', data) as mes,
                   SUM(valor) as total
            FROM receitas
            WHERE user_id=? AND date(data) >= date('now', '-6 months')
            GROUP BY mes
            ORDER BY mes
        '''
        receitas_mes = db_execute(conn, query, (user_id,)).fetchall() or []
    except:
        receitas_mes = []
    
    try:
        query = '''
            SELECT strftime('%Y-%m', data) as mes,
                   SUM(valor) as total
            FROM gastos
            WHERE user_id=? AND date(data) >= date('now', '-6 months')
            GROUP BY mes
            ORDER BY mes
        '''
        gastos_mes = db_execute(conn, query, (user_id,)).fetchall() or []
    except:
        gastos_mes = []
    
    conn.close()
    
    return jsonify({
        'receitas': [{'mes': r['mes'], 'valor': float(r['total'] or 0)} for r in receitas_mes],
        'gastos': [{'mes': g['mes'], 'valor': float(g['total'] or 0)} for g in gastos_mes]
    })

@app.route('/api/metas', methods=['GET', 'POST', 'PUT', 'DELETE'])
@login_required
def metas():
    conn = get_db()
    user_id = session.get('user_id')
    
    if request.method == 'POST':
        data = request.json
        db_execute(conn, '''INSERT INTO metas (user_id, titulo, valor_alvo, valor_atual, data_inicio, data_fim, tipo)
                       VALUES (?, ?, ?, ?, ?, ?, ?)''',
                    (user_id, data['titulo'], data['valor_alvo'], data.get('valor_atual', 0),
                     data['data_inicio'], data['data_fim'], data['tipo']))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    
    elif request.method == 'PUT':
        data = request.json
        db_execute(conn, 'UPDATE metas SET valor_atual = ? WHERE id = ? AND user_id = ?',
                    (data['valor_atual'], data['id'], user_id))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    
    elif request.method == 'DELETE':
        meta_id = request.args.get('id')
        db_execute(conn, 'DELETE FROM metas WHERE id = ? AND user_id = ?', (meta_id, user_id))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    
    metas = db_execute(conn, 'SELECT * FROM metas WHERE user_id=? AND ativo = 1 ORDER BY data_fim', (user_id,)).fetchall()
    conn.close()
    return jsonify([dict(m) for m in metas])

@app.route('/api/alertas')
@api_error_handler
@login_required
def alertas():
    conn = get_db()
    user_id = session.get('user_id')
    alertas = db_execute(conn, 'SELECT * FROM alertas WHERE user_id=? AND lido = 0 ORDER BY data DESC LIMIT 10', (user_id,)).fetchall()
    conn.close()
    return jsonify([dict(a) for a in alertas])

@app.route('/api/alertas/marcar-lido/<int:alerta_id>', methods=['POST'])
@login_required
def marcar_alerta_lido(alerta_id):
    conn = get_db()
    user_id = session.get('user_id')
    db_execute(conn, 'UPDATE alertas SET lido = 1 WHERE id = ? AND user_id = ?', (alerta_id, user_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/comparacao')
@api_error_handler
@login_required
def comparacao():
    conn = get_db()
    user_id = session.get('user_id')
    
    # Mês atual
    try:
        mes_atual_result = db_execute(conn, '''
            SELECT SUM(valor) as total FROM gastos 
            WHERE user_id=? AND strftime('%Y-%m', data) = strftime('%Y-%m', 'now')
        ''', (user_id,)).fetchone()
        mes_atual = float(mes_atual_result['total']) if mes_atual_result and mes_atual_result['total'] else 0
    except:
        mes_atual = 0
    
    # Mês anterior
    try:
        mes_anterior_result = db_execute(conn, '''
            SELECT SUM(valor) as total FROM gastos 
            WHERE user_id=? AND strftime('%Y-%m', data) = strftime('%Y-%m', date('now', '-1 month'))
        ''', (user_id,)).fetchone()
        mes_anterior = float(mes_anterior_result['total']) if mes_anterior_result and mes_anterior_result['total'] else 0
    except:
        mes_anterior = 0
    
    # Média dos últimos 3 meses
    try:
        media_result = db_execute(conn, '''
            SELECT AVG(total) as media FROM (
                SELECT SUM(valor) as total FROM gastos 
                WHERE user_id=? AND date(data) >= date('now', '-3 months')
                GROUP BY strftime('%Y-%m', data)
            )
        ''', (user_id,)).fetchone()
        media_3_meses = float(media_result['media']) if media_result and media_result['media'] else 0
    except:
        media_3_meses = 0
    
    conn.close()
    
    variacao = 0
    if mes_anterior > 0:
        variacao = ((mes_atual - mes_anterior) / mes_anterior) * 100
    
    return jsonify({
        'mes_atual': mes_atual,
        'mes_anterior': mes_anterior,
        'media_3_meses': media_3_meses,
        'variacao_percentual': variacao
    })

@app.route('/api/backup')
@api_error_handler
@login_required
def backup():
    import json
    from datetime import datetime
    
    conn = get_db()
    user_id = session.get('user_id')
    receitas = db_execute(conn, 'SELECT * FROM receitas WHERE user_id=?', (user_id,)).fetchall()
    gastos = db_execute(conn, 'SELECT * FROM gastos WHERE user_id=?', (user_id,)).fetchall()
    metas = db_execute(conn, 'SELECT * FROM metas WHERE user_id=?', (user_id,)).fetchall()
    conn.close()
    
    backup_data = {
        'data_backup': datetime.now().isoformat(),
        'receitas': [dict(r) for r in receitas],
        'gastos': [dict(g) for g in gastos],
        'metas': [dict(m) for m in metas]
    }
    
    return jsonify(backup_data)

# NOVAS ROTAS - FEATURES AVANÇADAS

@app.route('/api/comprovantes', methods=['GET', 'POST'])
@login_required
def comprovantes():
    from datetime import datetime
    conn = get_db()
    user_id = session.get('user_id')
    
    if request.method == 'POST':
        data = request.json
        db_execute(conn, '''INSERT INTO comprovantes (user_id, transacao_tipo, transacao_id, arquivo_base64, nome_arquivo, data_upload)
                       VALUES (?, ?, ?, ?, ?, ?)''',
                    (user_id, data['transacao_tipo'], data['transacao_id'], data['arquivo_base64'],
                     data.get('nome_arquivo', 'comprovante.jpg'), datetime.now().isoformat()))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    
    tipo = request.args.get('tipo')
    tid = request.args.get('id')
    comprovantes = db_execute(conn, 'SELECT * FROM comprovantes WHERE user_id=? AND transacao_tipo = ? AND transacao_id = ?',
                                (user_id, tipo, tid)).fetchall()
    conn.close()
    return jsonify([dict(c) for c in comprovantes])

@app.route('/api/recorrentes', methods=['GET', 'POST', 'PUT', 'DELETE'])
@login_required
def recorrentes():
    conn = get_db()
    user_id = session.get('user_id')
    
    if request.method == 'POST':
        data = request.json
        db_execute(conn, '''INSERT INTO recorrentes (user_id, descricao, valor, categoria, dia_vencimento)
                       VALUES (?, ?, ?, ?, ?)''',
                    (user_id, data['descricao'], data['valor'], data['categoria'], data['dia_vencimento']))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    
    elif request.method == 'PUT':
        data = request.json
        db_execute(conn, 'UPDATE recorrentes SET ativo = ? WHERE id = ? AND user_id = ?', (data['ativo'], data['id'], user_id))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    
    elif request.method == 'DELETE':
        rid = request.args.get('id')
        db_execute(conn, 'DELETE FROM recorrentes WHERE id = ? AND user_id = ?', (rid, user_id))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    
    recorrentes = db_execute(conn, 'SELECT * FROM recorrentes WHERE user_id=? AND ativo = 1', (user_id,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in recorrentes])

@app.route('/api/recorrentes/gerar', methods=['POST'])
@login_required
def gerar_recorrentes():
    from datetime import datetime
    conn = get_db()
    user_id = session.get('user_id')
    hoje = datetime.now()
    mes_atual = hoje.strftime('%Y-%m')
    
    recorrentes = db_execute(conn, 'SELECT * FROM recorrentes WHERE user_id=? AND ativo = 1', (user_id,)).fetchall()
    gerados = 0
    
    for rec in recorrentes:
        ultima = rec['ultima_geracao']
        if not ultima or ultima[:7] != mes_atual:
            dia = min(rec['dia_vencimento'], 28)
            data_gasto = f"{mes_atual}-{dia:02d}"
            
            db_execute(conn, 'INSERT INTO gastos (user_id, descricao, valor, categoria, data, tags) VALUES (?, ?, ?, ?, ?, ?)',
                        (user_id, rec['descricao'], rec['valor'], rec['categoria'], data_gasto, 'recorrente'))
            db_execute(conn, 'UPDATE recorrentes SET ultima_geracao = ? WHERE id = ? AND user_id = ?', (mes_atual, rec['id'], user_id))
            gerados += 1
    
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'gerados': gerados})

@app.route('/api/previsoes')
@api_error_handler
@login_required
def previsoes():
    from datetime import datetime
    conn = get_db()
    user_id = session.get('user_id')
    
    categorias = ['Alimentação', 'Transporte', 'Moradia', 'Saúde', 'Lazer', 'Educação', 'Outros']
    mes_atual = datetime.now().strftime('%Y-%m')
    previsoes = []
    
    for cat in categorias:
        try:
            media_result = db_execute(conn, '''
                SELECT AVG(total) as media FROM (
                    SELECT SUM(valor) as total FROM gastos 
                    WHERE user_id=? AND categoria = ? AND date(data) >= date('now', '-3 months')
                    GROUP BY strftime('%Y-%m', data)
                )
            ''', (user_id, cat)).fetchone()
            media = float(media_result['media']) if media_result and media_result['media'] else 0
        except:
            media = 0
        
        try:
            real_result = db_execute(conn, '''
                SELECT SUM(valor) as total FROM gastos 
                WHERE user_id=? AND categoria = ? AND strftime('%Y-%m', data) = ?
            ''', (user_id, cat, mes_atual)).fetchone()
            real = float(real_result['total']) if real_result and real_result['total'] else 0
        except:
            real = 0
        
        previsoes.append({
            'categoria': cat,
            'previsto': round(media, 2),
            'real': round(real, 2),
            'percentual': round((real / media * 100) if media > 0 else 0, 1)
        })
    
    conn.close()
    return jsonify(previsoes)
    return jsonify(previsoes)

@app.route('/api/gastos/mes')
@api_error_handler
@login_required
def gastos_por_mes():
    from datetime import datetime
    conn = get_db()
    user_id = session.get('user_id')
    mes = request.args.get('mes', datetime.now().strftime('%Y-%m'))
    
    gastos = db_execute(conn, '''
        SELECT *, 
        CASE WHEN tags LIKE '%recorrente%' THEN 1 ELSE 0 END as eh_recorrente
        FROM gastos 
        WHERE user_id=? AND strftime('%Y-%m', data) = ?
        ORDER BY eh_recorrente DESC, data DESC
    ''', (user_id, mes)).fetchall()
    
    conn.close()
    return jsonify([dict(g) for g in gastos])

@app.route('/api/relatorio-ir')
@api_error_handler
@login_required
def relatorio_ir():
    from datetime import datetime
    conn = get_db()
    user_id = session.get('user_id')
    ano = request.args.get('ano', datetime.now().year)
    
    try:
        receitas_result = db_execute(conn, '''
            SELECT SUM(valor) as total FROM receitas 
            WHERE user_id=? AND strftime('%Y', data) = ? AND tipo IN ('Salário', 'Freelance')
        ''', (user_id, str(ano))).fetchone()
        receitas_tributaveis = float(receitas_result['total']) if receitas_result and receitas_result['total'] else 0
    except:
        receitas_tributaveis = 0
    
    try:
        despesas_dedutiveis = db_execute(conn, '''
            SELECT categoria, SUM(valor) as total FROM gastos 
            WHERE user_id=? AND strftime('%Y', data) = ? AND categoria IN ('Saúde', 'Educação')
            GROUP BY categoria
        ''', (user_id, str(ano))).fetchall() or []
    except:
        despesas_dedutiveis = []
    
    conn.close()
    
    return jsonify({
        'ano': ano,
        'receitas_tributaveis': receitas_tributaveis,
        'despesas_dedutiveis': [dict(d) for d in despesas_dedutiveis],
        'total_deducoes': sum(float(d['total'] or 0) for d in despesas_dedutiveis)
    })

@app.route('/api/tags')
@api_error_handler
@login_required
def listar_tags():
    conn = get_db()
    user_id = session.get('user_id')
    tags_gastos = db_execute(conn, 'SELECT DISTINCT tags FROM gastos WHERE user_id=? AND tags IS NOT NULL AND tags != ""', (user_id,)).fetchall()
    tags_receitas = db_execute(conn, 'SELECT DISTINCT tags FROM receitas WHERE user_id=? AND tags IS NOT NULL AND tags != ""', (user_id,)).fetchall()
    conn.close()
    
    todas_tags = set()
    for t in (tags_gastos or []) + (tags_receitas or []):
        if t and t.get('tags'):
            todas_tags.update(t['tags'].split(','))
    
    return jsonify(sorted([t.strip() for t in todas_tags if t and t.strip()]))

@app.route('/api/exportar/excel')
@api_error_handler
@login_required
def exportar_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from flask import send_file
    
    conn = get_db()
    user_id = session.get('user_id')
    receitas = db_execute(conn, 'SELECT * FROM receitas WHERE user_id=? ORDER BY data DESC', (user_id,)).fetchall()
    gastos = db_execute(conn, 'SELECT * FROM gastos WHERE user_id=? ORDER BY data DESC', (user_id,)).fetchall()
    metas = db_execute(conn, 'SELECT * FROM metas WHERE user_id=? ORDER BY data_inicio DESC', (user_id,)).fetchall()
    conn.close()
    
    wb = Workbook()
    
    # Aba Receitas
    ws_receitas = wb.active
    ws_receitas.title = "Receitas"
    ws_receitas.append(['Data', 'Descrição', 'Valor', 'Tipo', 'Notas', 'Tags'])
    for r in receitas:
        ws_receitas.append([r['data'], r['descricao'], r['valor'], r['tipo'], r['notas'], r['tags']])
    
    # Aba Gastos
    ws_gastos = wb.create_sheet("Gastos")
    ws_gastos.append(['Data', 'Descrição', 'Valor', 'Categoria', 'Notas', 'Tags'])
    for g in gastos:
        ws_gastos.append([g['data'], g['descricao'], g['valor'], g['categoria'], g['notas'], g['tags']])
    
    # Aba Metas
    ws_metas = wb.create_sheet("Metas")
    ws_metas.append(['Título', 'Valor Alvo', 'Valor Atual', 'Data Início', 'Data Fim', 'Tipo', 'Ativo'])
    for m in metas:
        ws_metas.append([m['titulo'], m['valor_alvo'], m['valor_atual'], m['data_inicio'], m['data_fim'], m['tipo'], 'Sim' if m['ativo'] else 'Não'])
    
    # Estilizar cabeçalhos
    for ws in [ws_receitas, ws_gastos, ws_metas]:
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'financeiro_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

@app.route('/api/exportar/modelo-excel')
def exportar_modelo_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from flask import send_file
    from datetime import datetime
    
    wb = Workbook()
    
    # Aba Receitas
    ws_receitas = wb.active
    ws_receitas.title = "Receitas"
    ws_receitas.append(['Data', 'Descrição', 'Valor', 'Tipo', 'Notas', 'Tags'])
    ws_receitas.append(['2026-02-25', 'Salário', 5000.00, 'Salário', 'Pagamento mensal', 'trabalho,fixo'])
    ws_receitas.append(['', '', '', 'Salário/Freelance/Investimento/Outros', '', ''])
    
    # Aba Gastos
    ws_gastos = wb.create_sheet("Gastos")
    ws_gastos.append(['Data', 'Descrição', 'Valor', 'Categoria', 'Notas', 'Tags'])
    ws_gastos.append(['2026-02-25', 'Aluguel', 1200.00, 'Moradia', 'Apartamento', 'fixo,mensal'])
    ws_gastos.append(['', '', '', 'Alimentação/Transporte/Moradia/Saúde/Lazer/Educação/Outros', '', ''])
    
    # Aba Metas
    ws_metas = wb.create_sheet("Metas")
    ws_metas.append(['Título', 'Valor Alvo', 'Valor Atual', 'Data Início', 'Data Fim', 'Tipo', 'Ativo'])
    ws_metas.append(['Viagem', 5000.00, 1500.00, '2026-01-01', '2026-12-31', 'Lazer', 'Sim'])
    ws_metas.append(['', '', '', '', '', 'Economia/Investimento/Lazer/Outros', 'Sim/Não'])
    
    # Aba Instruções
    ws_instrucoes = wb.create_sheet("📋 INSTRUÇÕES")
    instrucoes = [
        ['COMO USAR ESTA PLANILHA'],
        [''],
        ['1. FORMATO DAS DATAS'],
        ['   Use o formato: AAAA-MM-DD (exemplo: 2026-02-25)'],
        [''],
        ['2. VALORES'],
        ['   Use ponto para decimais (exemplo: 1500.50)'],
        ['   Não use vírgulas ou símbolos de moeda'],
        [''],
        ['3. TIPOS E CATEGORIAS'],
        ['   Receitas: Salário, Freelance, Investimento, Outros'],
        ['   Gastos: Alimentação, Transporte, Moradia, Saúde, Lazer, Educação, Outros'],
        ['   Metas: Economia, Investimento, Lazer, Outros'],
        [''],
        ['4. TAGS (OPCIONAL)'],
        ['   Separe múltiplas tags com vírgula (exemplo: fixo,mensal)'],
        [''],
        ['5. METAS - CAMPO ATIVO'],
        ['   Use "Sim" para ativa ou "Não" para inativa'],
        [''],
        ['6. LINHA DE EXEMPLO'],
        ['   A primeira linha de cada aba tem um exemplo preenchido'],
        ['   A segunda linha mostra as opções válidas'],
        ['   Você pode apagar essas linhas antes de importar'],
        [''],
        ['7. IMPORTAR'],
        ['   Preencha os dados nas abas Receitas, Gastos e Metas'],
        ['   Salve o arquivo'],
        ['   No sistema, clique em "📂 Importar Excel"'],
        ['   Selecione este arquivo'],
        [''],
        ['⚠️ IMPORTANTE:'],
        ['   - Não altere os nomes das colunas (cabeçalhos)'],
        ['   - Não altere os nomes das abas'],
        ['   - Mantenha a ordem das colunas'],
        ['   - Linhas vazias serão ignoradas'],
    ]
    
    for row in instrucoes:
        ws_instrucoes.append(row)
    
    # Estilizar cabeçalhos
    for ws in [ws_receitas, ws_gastos, ws_metas]:
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        # Linha de exemplo em amarelo
        for cell in ws[2]:
            cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        # Linha de instruções em cinza claro
        for cell in ws[3]:
            cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            cell.font = Font(italic=True, size=9)
    
    # Estilizar instruções
    ws_instrucoes['A1'].font = Font(bold=True, size=14, color="667eea")
    for row in ws_instrucoes.iter_rows(min_row=3):
        if row[0].value and row[0].value.startswith('   '):
            row[0].font = Font(size=10)
        elif row[0].value and any(x in str(row[0].value) for x in ['1.', '2.', '3.', '4.', '5.', '6.', '7.']):
            row[0].font = Font(bold=True, size=11)
        elif row[0].value and '⚠️' in str(row[0].value):
            row[0].font = Font(bold=True, size=11, color="FF0000")
    
    # Ajustar largura das colunas
    for ws in [ws_receitas, ws_gastos, ws_metas]:
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 20
    
    ws_instrucoes.column_dimensions['A'].width = 80
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='modelo_importacao.xlsx'
    )

@app.route('/api/importar/excel', methods=['POST'])
@login_required
def importar_excel():
    from openpyxl import load_workbook
    from io import BytesIO
    
    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Arquivo vazio'}), 400
    
    try:
        wb = load_workbook(BytesIO(file.read()))
        conn = get_db()
        user_id = session.get('user_id')
        
        importados = {'receitas': 0, 'gastos': 0, 'metas': 0}
        
        # Importar Receitas
        if 'Receitas' in wb.sheetnames:
            ws = wb['Receitas']
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Se tem data
                    db_execute(conn, '''INSERT INTO receitas (user_id, data, descricao, valor, tipo, notas, tags)
                                    VALUES (?, ?, ?, ?, ?, ?, ?)''',
                                (user_id, row[0], row[1], row[2], row[3], row[4] or '', row[5] or ''))
                    importados['receitas'] += 1
        
        # Importar Gastos
        if 'Gastos' in wb.sheetnames:
            ws = wb['Gastos']
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    db_execute(conn, '''INSERT INTO gastos (user_id, data, descricao, valor, categoria, notas, tags)
                                    VALUES (?, ?, ?, ?, ?, ?, ?)''',
                                (user_id, row[0], row[1], row[2], row[3], row[4] or '', row[5] or ''))
                    importados['gastos'] += 1
        
        # Importar Metas
        if 'Metas' in wb.sheetnames:
            ws = wb['Metas']
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    ativo = 1 if row[6] == 'Sim' else 0
                    db_execute(conn, '''INSERT INTO metas (user_id, titulo, valor_alvo, valor_atual, data_inicio, data_fim, tipo, ativo)
                                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                                (user_id, row[0], row[1], row[2], row[3], row[4], row[5], ativo))
                    importados['metas'] += 1
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'Importado: {importados["receitas"]} receitas, {importados["gastos"]} gastos, {importados["metas"]} metas'
        })
    
    except Exception as e:
        return jsonify({'error': f'Erro ao importar: {str(e)}'}), 500

# Proteção contra CSRF (básica)
@app.route('/api/migrar-db')
def migrar_db():
    """Rota para forçar migração do banco (criar tabelas)"""
    try:
        from database import init_db
        init_db()
        return jsonify({'success': True, 'message': 'Banco de dados migrado com sucesso!'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/adicionar-user-id')
def adicionar_user_id():
    """Adiciona coluna user_id em todas as tabelas que não têm"""
    try:
        conn = get_db()
        is_postgres = hasattr(conn, 'cursor_factory')
        cursor = conn.cursor()
        
        tabelas = ['receitas', 'gastos', 'metas', 'alertas', 'comprovantes', 'recorrentes', 'previsoes']
        
        for tabela in tabelas:
            try:
                # Tenta adicionar a coluna (ignora se já existe)
                if is_postgres:
                    cursor.execute(f'ALTER TABLE {tabela} ADD COLUMN IF NOT EXISTS user_id INTEGER DEFAULT 1')
                else:
                    # SQLite não tem IF NOT EXISTS no ALTER TABLE
                    try:
                        cursor.execute(f'ALTER TABLE {tabela} ADD COLUMN user_id INTEGER DEFAULT 1')
                    except:
                        pass  # Coluna já existe
                conn.commit()
                print(f"✅ Coluna user_id adicionada em {tabela}")
            except Exception as e:
                print(f"⚠️ Erro em {tabela}: {e}")
        
        conn.close()
        return jsonify({'success': True, 'message': 'Colunas user_id adicionadas com sucesso!'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/test-sqlalchemy')
def test_sqlalchemy():
    """Rota de teste para verificar se SQLAlchemy está funcionando"""
    try:
        from models.user import User
        from models.transaction import Receita, Gasto
        from models.meta import Meta
        
        # Tenta contar registros usando ORM
        user_count = User.query.count()
        receita_count = Receita.query.count()
        gasto_count = Gasto.query.count()
        meta_count = Meta.query.count()
        
        return jsonify({
            'success': True,
            'message': '✅ SQLAlchemy funcionando perfeitamente!',
            'orm_counts': {
                'usuarios': user_count,
                'receitas': receita_count,
                'gastos': gasto_count,
                'metas': meta_count
            },
            'database': app.config.get('SQLALCHEMY_DATABASE_URI', 'N/A')[:20] + '...'
        })
    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

@app.after_request
def set_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    return response

if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
